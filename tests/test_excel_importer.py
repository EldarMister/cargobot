from datetime import datetime

import openpyxl
import xlwt

from app.services.excel_importer import ExcelImporter


def test_xlsx_reads_all_sheets_and_skips_headers_and_dates(tmp_path):
    path = tmp_path / "cargo.xlsx"
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "Sheet1"
    first.append(["Трек-код", "Код клиента"])
    first.append(["YT7592444294461", "h-829"])
    first.append([datetime(2025, 12, 22), None])
    first.append(["AD2025年12月22日", None])
    second = workbook.create_sheet("Sheet2")
    second.append([9812328869266, "H-4040"])
    workbook.save(path)

    result = ExcelImporter().parse(path)

    assert [(row.sheet_name, row.tracking_number, row.client_code) for row in result.valid_rows] == [
        ("Sheet1", "YT7592444294461", "H-829"),
        ("Sheet2", "9812328869266", "H-4040"),
    ]
    assert any("код клиента" in row.error.lower() for row in result.skipped_rows)
    assert any("дат" in row.error.lower() for row in result.skipped_rows)


def test_xls_is_supported(tmp_path):
    path = tmp_path / "cargo.xls"
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("OldFormat")
    sheet.write(0, 0, "Tracking")
    sheet.write(0, 1, "Client")
    sheet.write(1, 0, "78999695208956")
    sheet.write(1, 1, "H-8226")
    workbook.save(str(path))

    result = ExcelImporter().parse(path)

    assert len(result.valid_rows) == 1
    assert result.valid_rows[0].tracking_number == "78999695208956"
    assert result.valid_rows[0].client_code == "H-8226"


def test_invalid_extension_is_rejected(tmp_path):
    path = tmp_path / "cargo.csv"
    path.write_text("track,code", encoding="utf-8")
    try:
        ExcelImporter().parse(path)
    except ValueError as exc:
        assert ".xls" in str(exc)
    else:
        raise AssertionError("Unsupported extension was accepted")
