from __future__ import annotations

from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import openpyxl
import xlrd

from app.services.normalization import (
    is_valid_client_code,
    is_valid_tracking_number,
    normalize_client_code,
    normalize_tracking_number,
)


@dataclass(slots=True)
class ParsedExcelRow:
    sheet_name: str
    row_number: int
    tracking_number: str | None = None
    client_code: str | None = None
    error: str | None = None

    @property
    def is_valid(self) -> bool:
        return self.error is None and bool(self.tracking_number and self.client_code)


@dataclass(slots=True)
class ExcelParseResult:
    rows: list[ParsedExcelRow] = field(default_factory=list)

    @property
    def total_rows(self) -> int:
        return len(self.rows)

    @property
    def valid_rows(self) -> list[ParsedExcelRow]:
        return [row for row in self.rows if row.is_valid]

    @property
    def skipped_rows(self) -> list[ParsedExcelRow]:
        return [row for row in self.rows if not row.is_valid]


class ExcelImporter:
    """Reads every sheet and extracts tracking/client-code pairs without Telegram dependencies."""

    def parse(self, path: str | Path) -> ExcelParseResult:
        path = Path(path)
        suffix = path.suffix.lower()
        if suffix == ".xlsx":
            sheets = self._read_xlsx(path)
        elif suffix == ".xls":
            sheets = self._read_xls(path)
        else:
            raise ValueError("Поддерживаются только файлы .xls и .xlsx")

        result = ExcelParseResult()
        for sheet_name, rows in sheets:
            for row_number, values in rows:
                if not any(value not in (None, "") for value in values):
                    continue
                result.rows.append(self._parse_row(sheet_name, row_number, values))
        return result

    @staticmethod
    def _read_xlsx(path: Path):
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in workbook.worksheets:
                yield (
                    sheet.title,
                    (
                        (index, tuple(cell.value for cell in row))
                        for index, row in enumerate(sheet.iter_rows(), start=1)
                    ),
                )
        finally:
            workbook.close()

    @staticmethod
    def _read_xls(path: Path):
        workbook = xlrd.open_workbook(path, on_demand=True)
        try:
            for sheet in workbook.sheets():

                def iter_rows(current_sheet=sheet):
                    for row_index in range(current_sheet.nrows):
                        values: list[object] = []
                        for cell in current_sheet.row(row_index):
                            if cell.ctype == xlrd.XL_CELL_DATE:
                                values.append(xlrd.xldate_as_datetime(cell.value, workbook.datemode))
                            elif cell.ctype == xlrd.XL_CELL_NUMBER and float(cell.value).is_integer():
                                values.append(int(cell.value))
                            else:
                                values.append(cell.value)
                        yield row_index + 1, tuple(values)

                yield sheet.name, iter_rows()
        finally:
            workbook.release_resources()

    @staticmethod
    def _parse_row(sheet_name: str, row_number: int, values: Iterable[object]) -> ParsedExcelRow:
        values = tuple(values)
        if any(isinstance(value, (date, datetime)) for value in values):
            return ParsedExcelRow(sheet_name, row_number, error="Строка содержит дату")

        client_codes = [normalize_client_code(value) for value in values if is_valid_client_code(value)]
        tracks = [normalize_tracking_number(value) for value in values if is_valid_tracking_number(value)]
        tracks = [track for track in tracks if track not in client_codes]

        if not client_codes:
            return ParsedExcelRow(sheet_name, row_number, error="Не найден корректный код клиента")
        if not tracks:
            return ParsedExcelRow(
                sheet_name, row_number, client_code=client_codes[0], error="Не найден корректный трек-код"
            )
        return ParsedExcelRow(
            sheet_name=sheet_name,
            row_number=row_number,
            tracking_number=tracks[0],
            client_code=client_codes[0],
        )
